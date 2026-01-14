import json
import os
from pathlib import Path
import openpyxl

def read_json_rules(file_path):
    """Read rules from JSON file"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            if isinstance(data, list):
                return data
            elif isinstance(data, dict) and 'rules' in data:
                return data['rules']
            else:
                return [str(data)]
    except Exception as e:
        print(f"Error reading JSON {file_path}: {e}")
        return []

def read_txt_rules(file_path):
    """Read rules from TXT file"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return [line.strip() for line in f if line.strip()]
    except Exception as e:
        print(f"Error reading TXT {file_path}: {e}")
        return []

def read_xlsx_rules(file_path):
    """Read rules from XLSX file"""
    try:
        rules = []
        wb = openpyxl.load_workbook(file_path)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell and str(cell).strip():
                        rules.append(str(cell).strip())
        return rules
    except Exception as e:
        print(f"Error reading XLSX {file_path}: {e}")
        return []

def generate_all_rules(rules_dir="./审核规则库"):
    """Combine all rules from JSON, TXT, XLSX files"""
    all_rules = []
    rules_path = Path(rules_dir)
    
    if not rules_path.exists():
        print(f"Directory {rules_dir} not found")
        return
    
    # Process JSON files
    for json_file in rules_path.glob("*.json"):
        all_rules.extend(read_json_rules(json_file))
    
    # Process TXT files
    for txt_file in rules_path.glob("*.txt"):
        all_rules.extend(read_txt_rules(txt_file))
    
    # Process XLSX files
    for xlsx_file in rules_path.glob("*.xlsx"):
        all_rules.extend(read_xlsx_rules(xlsx_file))
    
    # Convert all rules to strings and remove duplicates
    all_rules = [str(rule) for rule in all_rules]
    all_rules = list(set(all_rules))
    
    output_file = rules_path / "all_rules.txt"
    with open(output_file, 'w', encoding='utf-8') as f:
        for rule in sorted(all_rules):
            f.write(rule + '\n')
    
    print(f"Successfully generated {output_file}")
    print(f"Total rules: {len(all_rules)}")

if __name__ == "__main__":
    generate_all_rules()